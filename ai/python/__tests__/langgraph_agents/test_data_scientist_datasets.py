import sys
from pathlib import Path

import openpyxl
import pytest

PYTHON_ROOT = Path(__file__).resolve().parents[2]
AI_ROOT = PYTHON_ROOT.parent
for path in (str(PYTHON_ROOT), str(AI_ROOT)):
    if path not in sys.path:
        sys.path.append(path)

from backend.agents.data_scientist import dataset_io


def test_load_manifest_returns_empty_list_when_missing(tmp_path):
    assert dataset_io.load_manifest(tmp_path) == []


def test_load_manifest_reads_json_manifest(tmp_path):
    manifest = tmp_path / "datasets.json"
    manifest.write_text('[{"id":"d1"}]', encoding="utf-8")
    assert dataset_io.load_manifest(tmp_path) == [{"id": "d1"}]


def test_resolve_dataset_entry_returns_matching_record():
    manifest = [{"id": "a"}, {"id": "b", "name": "Dataset B"}]
    assert dataset_io.resolve_dataset_entry("b", manifest) == {"id": "b", "name": "Dataset B"}
    assert dataset_io.resolve_dataset_entry("c", manifest) is None


def test_clean_header_strips_bom_and_whitespace():
    assert dataset_io._clean_header(" \ufefffeature ") == "feature"


def test_row_mapping_skips_blank_headers():
    row = dataset_io._row_mapping(["feature", "", "target"], [1, "skip", "yes"])
    assert row == {"feature": 1, "target": "yes"}


def test_detect_delimiter_prefers_semicolon(tmp_path):
    path = tmp_path / "semi.csv"
    path.write_text("a;b\n1;2\n", encoding="utf-8")
    with path.open("r", encoding="utf-8") as handle:
        assert dataset_io.detect_delimiter(handle) == ";"


def test_detect_delimiter_defaults_to_comma(tmp_path):
    path = tmp_path / "plain.txt"
    path.write_text("ab\n12\n", encoding="utf-8")
    with path.open("r", encoding="utf-8") as handle:
        assert dataset_io.detect_delimiter(handle) == ","


def test_load_dataset_rows_dispatches_with_custom_row_loaders(tmp_path):
    xlsx_path = tmp_path / "sheet.xlsx"
    xls_path = tmp_path / "sheet.xls"

    rows_xlsx = dataset_io.load_dataset_rows(xlsx_path, row_loaders={".xlsx": lambda _path: [{"kind": "xlsx"}]})
    rows_xls = dataset_io.load_dataset_rows(xls_path, row_loaders={".xls": lambda _path: [{"kind": "xls"}]})

    assert rows_xlsx == [{"kind": "xlsx"}]
    assert rows_xls == [{"kind": "xls"}]


def test_read_xlsx_rows_reads_first_sheet(tmp_path):
    path = tmp_path / "dataset.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["feature", "target"])
    sheet.append([1, "yes"])
    sheet.append([2, "no"])
    workbook.save(path)

    rows = dataset_io.read_xlsx_rows(path)
    assert rows == [{"feature": 1, "target": "yes"}, {"feature": 2, "target": "no"}]


def test_read_csv_rows_normalizes_headers_and_detects_tab_delimiter(tmp_path):
    path = tmp_path / "dataset.csv"
    path.write_text("\ufefffeature\t target \n1\tyes\n2\tno\n", encoding="utf-8")
    rows = dataset_io.read_csv_rows(path)
    assert rows == [{"feature": "1", "target": "yes"}, {"feature": "2", "target": "no"}]


def test_read_xlsx_rows_returns_empty_without_headers(tmp_path):
    path = tmp_path / "empty.xlsx"
    workbook = openpyxl.Workbook()
    workbook.active.append([])
    workbook.save(path)
    assert dataset_io.read_xlsx_rows(path) == []


def test_read_xls_rows_reads_first_sheet(monkeypatch, tmp_path):
    class _Sheet:
        nrows = 3

        def row_values(self, index):
            if index == 0:
                return ["feature", "", "target"]
            if index == 1:
                return [1.0, "skip", "yes"]
            return [2.0, "skip", "no"]

    class _Workbook:
        def sheet_by_index(self, index):
            assert index == 0
            return _Sheet()

    monkeypatch.setattr(dataset_io.xlrd, "open_workbook", lambda path: _Workbook())
    rows = dataset_io.read_xls_rows(tmp_path / "dataset.xls")
    assert rows == [{"feature": 1.0, "target": "yes"}, {"feature": 2.0, "target": "no"}]


def test_load_dataset_rows_rejects_unknown_suffix(tmp_path):
    with pytest.raises(ValueError, match="Unsupported file format"):
        dataset_io.load_dataset_rows(tmp_path / "dataset.json")
