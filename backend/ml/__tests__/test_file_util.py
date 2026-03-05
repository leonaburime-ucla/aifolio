import math
import sys
from pathlib import Path

import openpyxl
import pytest

AI_ROOT = Path(__file__).resolve().parents[2]
if str(AI_ROOT) not in sys.path:
    sys.path.append(str(AI_ROOT))

from ml.file_util import coerce_value, load_tabular_file, split_features_target
import ml.file_util as file_util


def test_sniff_csv_dialect_falls_back_to_excel_for_ambiguous_sample():
    dialect = file_util._sniff_csv_dialect("plain text without separators")
    assert dialect.delimiter == ","


def test_normalize_header_row_and_generated_header_detection():
    assert file_util._normalize_header_row([" X1 ", None, "Y"]) == ["X1", "", "Y"]
    assert file_util._looks_like_generated_header(["X1", "X2", "Y"]) is True
    assert file_util._looks_like_generated_header(["feature", "target"]) is False


def test_row_to_entry_skips_blank_columns_and_pads_missing_values():
    row = file_util._row_to_entry(["a", "", "c"], [1])
    assert row == {"a": 1, "c": None}


def test_load_tabular_file_reads_csv(tmp_path):
    path = tmp_path / "sample.csv"
    path.write_text("a,b,c\n1,2,3\n4,5,6\n", encoding="utf-8")
    rows = load_tabular_file(path)
    assert rows[0]["a"] == "1"
    assert rows[1]["c"] == "6"


def test_load_tabular_file_reads_xlsx(tmp_path):
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["x", "y"])
    ws.append([1, 2])
    wb.save(path)

    rows = load_tabular_file(path)
    assert rows == [{"x": 1, "y": 2}]


def test_load_tabular_file_reads_xlsx_after_generated_header_row(tmp_path):
    path = tmp_path / "generated.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["X1", "X2", "Y"])
    ws.append(["feature", "score", "target"])
    ws.append([1, 2, "yes"])
    wb.save(path)

    rows = load_tabular_file(path)
    assert rows == [{"feature": 1, "score": 2, "target": "yes"}]


def test_load_tabular_file_reads_specific_sheet_name(tmp_path):
    path = tmp_path / "workbook.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    ws = wb.create_sheet("Data")
    ws.append(["a", "b"])
    ws.append([1, 2])
    wb.save(path)

    rows = load_tabular_file(path, sheet_name="Data")
    assert rows == [{"a": 1, "b": 2}]


def test_read_xls_raises_without_xlrd(monkeypatch, tmp_path):
    monkeypatch.setattr(file_util, "xlrd", None)
    with pytest.raises(RuntimeError, match="xlrd is required"):
        file_util._read_xls(tmp_path / "data.xls")


def test_read_xls_uses_second_row_when_first_row_is_generated_header(monkeypatch, tmp_path):
    class _Sheet:
        nrows = 3
        ncols = 3

        def cell_value(self, row, col):
            rows = [
                ["X1", "X2", "Y"],
                ["feature", "score", "target"],
                [1.0, 2.0, "yes"],
            ]
            return rows[row][col]

    class _Workbook:
        def sheet_by_index(self, index):
            assert index == 0
            return _Sheet()

        def sheet_by_name(self, name):
            assert name == "Data"
            return _Sheet()

    class _XlrdModule:
        def open_workbook(self, path):
            return _Workbook()

    monkeypatch.setattr(file_util, "xlrd", _XlrdModule())
    rows = file_util._read_xls(tmp_path / "data.xls", sheet_name="Data")
    assert rows == [{"feature": 1.0, "score": 2.0, "target": "yes"}]


def test_load_tabular_file_rejects_unsupported_extension(tmp_path):
    with pytest.raises(ValueError, match="Unsupported file extension"):
        load_tabular_file(tmp_path / "sample.json")


def test_coerce_value_maps_missing_to_nan_and_numeric_text_to_float():
    assert math.isnan(coerce_value(None))
    assert coerce_value(7) == 7
    assert math.isnan(coerce_value("NA"))
    assert math.isnan(coerce_value("."))
    assert coerce_value("3.5") == 3.5
    assert coerce_value("category") == "category"


def test_split_features_target_validates_input():
    with pytest.raises(ValueError):
        split_features_target([], "y")


def test_split_features_target_rejects_missing_target_column():
    with pytest.raises(ValueError, match="Target column 'y' not found"):
        split_features_target([{"x": 1}], "y")


def test_split_features_target_skips_rows_missing_target_and_coerces_features():
    x_rows, y = split_features_target(
        [
            {"feature": "3.5", "category": "A", "target": "1"},
            {"feature": "2.0"},
        ],
        "target",
    )
    assert y == ["1"]
    assert x_rows == [{"feature": 3.5, "category": "A"}]
